"""
Excel utilities untuk import/export data
Digunakan di semua app untuk konsistensi
"""
import csv
import io
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.db import models
from django.core.exceptions import ValidationError


class ExcelExporter:
    """Export data Django model ke Excel file"""
    
    def __init__(self, model, queryset=None, columns=None, title=None):
        """
        Args:
            model: Django model class
            queryset: QuerySet untuk di-export (default: semua data)
            columns: List field names [(field_name, display_name), ...]
                    Jika None, ambil dari model fields
            title: Judul sheet (default: model name)
        """
        self.model = model
        self.queryset = queryset or model.objects.all()
        self.title = title or model._meta.verbose_name_plural
        self.columns = columns or self._get_default_columns()
    
    def _get_default_columns(self):
        """Get semua field dari model kecuali ManyToMany relations"""
        columns = []
        for field in self.model._meta.get_fields():
            # Skip many-to-many dan one-to-many relations
            if field.one_to_many or field.many_to_many:
                continue

            # Primary key dibuat otomatis oleh sistem, jadi tidak ikut export
            if getattr(field, "primary_key", False):
                continue

            # Field audit sistem tidak perlu ikut template export/import default
            if getattr(field, "auto_now", False) or getattr(field, "auto_now_add", False):
                continue
            
            # Untuk FK, gunakan field_id agar bisa di-import dengan ID
            if field.many_to_one:
                field_name = f"{field.name}_id"
                label = f"{getattr(field, 'verbose_name', field.name).title()} (ID)"
            else:
                field_name = field.name
                label = getattr(field, 'verbose_name', field.name).title()
            
            columns.append((field_name, label))
        return columns
    
    def export(self):
        """Export ke Excel dan return file bytes"""
        wb = Workbook()
        ws = wb.active
        ws.title = self.title[:31]  # Sheet name max 31 chars
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_idx, (field_name, display_name) in enumerate(self.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = display_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        for row_idx, obj in enumerate(self.queryset, 2):
            for col_idx, (field_name, _) in enumerate(self.columns, 1):
                value = getattr(obj, field_name, None)
                
                # Format datetime
                if isinstance(value, datetime):
                    value = value.strftime('%d/%m/%Y %H:%M')
                elif isinstance(value, models.Model):
                    # Jangan sampai terjadi, karena sudah gunakan field_id
                    value = str(value)
                
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = data_alignment
                cell.border = border
        
        # Auto-adjust column widths
        for col_idx, (_, display_name) in enumerate(self.columns, 1):
            max_length = max(len(str(display_name)), 20)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
        
        # Return as bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class ExcelImporter:
    """Import & validasi data dari Excel file"""
    
    def __init__(
        self,
        model,
        file_stream,
        columns=None,
        skip_empty_rows=True,
        match_fields=None,
        filename=None,
    ):
        """
        Args:
            model: Django model untuk import
            file_stream: File object atau bytes
            columns: List field names untuk mapping dengan Excel columns
            skip_empty_rows: Skip baris kosong (default: True)
            match_fields: Field atau kombinasi field untuk menentukan update data existing
            filename: Nama file upload untuk deteksi format
        """
        self.model = model
        self.file_stream = file_stream
        self.columns = columns or self._get_default_columns()
        self.skip_empty_rows = skip_empty_rows
        self.match_fields = match_fields
        self.filename = filename
        self.errors = []
        self.read_errors = []
        self.empty_file_error = False
        self.preview_data = []
    
    def _get_default_columns(self):
        """Get semua field yang bisa di-import (kecuali many-to-many dan one-to-many)"""
        fields = []
        for field in self.model._meta.get_fields():
            # Skip many-to-many dan one-to-many relations
            if field.many_to_many or field.one_to_many:
                continue

            if getattr(field, "primary_key", False):
                continue

            if getattr(field, "auto_now", False) or getattr(field, "auto_now_add", False):
                continue
            
            # Untuk FK, gunakan field_id
            if field.many_to_one:
                fields.append(f"{field.name}_id")
            else:
                fields.append(field.name)
        return fields
    
    def _convert_field_id_to_field(self, field_name):
        """Convert field_id ke field name jika diperlukan"""
        if field_name.endswith('_id'):
            actual_field_name = field_name[:-3]  # Remove '_id'
            try:
                field = self.model._meta.get_field(actual_field_name)
                if field.many_to_one:
                    return actual_field_name, True  # field_name, is_fk=True
            except:
                pass
        return field_name, False

    def _is_empty_value(self, value):
        if value is None:
            return True
        if isinstance(value, str) and value.strip() == '':
            return True
        return False

    def _get_empty_field_value(self, field, is_fk_field=False):
        if is_fk_field or field.many_to_one or getattr(field, 'null', False):
            return None

        if isinstance(field, (models.CharField, models.TextField)) and getattr(field, 'blank', False):
            return ''

        return None

    def _has_matchable_value(self, field_name, value):
        if value is None:
            return False

        if value != '':
            return True

        field = self.model._meta.get_field(field_name)
        return isinstance(field, (models.CharField, models.TextField))

    def _prepare_processed_data(self, row_data):
        """Konversi data excel menjadi format yang siap divalidasi/disimpan"""
        processed_data = {}
        errors = []

        for field_col_name, value in row_data.items():
            field_name, is_fk_field = self._convert_field_id_to_field(field_col_name)
            field = self.model._meta.get_field(field_name)

            if self._is_empty_value(value):
                processed_data[field_name] = self._get_empty_field_value(field, is_fk_field)
                continue

            if is_fk_field or field.many_to_one:
                related_model = field.related_model

                try:
                    obj = self._resolve_related_instance(related_model, value)
                    processed_data[field_name] = obj
                except (related_model.DoesNotExist, ValueError):
                    errors.append(
                        f"Foreign key '{field_name}' dengan value '{value}' tidak ditemukan"
                    )
                    processed_data[field_name] = None

            elif field.__class__.__name__ == 'DateField' and value:
                try:
                    if isinstance(value, str):
                        processed_data[field_name] = datetime.strptime(str(value), '%d/%m/%Y').date()
                    else:
                        processed_data[field_name] = value
                except ValueError:
                    processed_data[field_name] = value
            else:
                processed_data[field_name] = value

        return processed_data, errors

    def _get_unique_fields(self):
        """Get forward fields yang punya unique/primary key"""
        unique_fields = []
        for field in self.model._meta.fields:
            if getattr(field, 'primary_key', False) or getattr(field, 'unique', False):
                unique_fields.append(field.name)
        return unique_fields

    def _get_unique_constraint_groups(self):
        groups = []

        for unique_group in getattr(self.model._meta, "unique_together", ()) or ():
            if unique_group:
                groups.append(tuple(unique_group))

        for constraint in getattr(self.model._meta, "constraints", ()):
            if not isinstance(constraint, models.UniqueConstraint):
                continue

            if getattr(constraint, "condition", None):
                continue

            if getattr(constraint, "expressions", ()):
                continue

            if getattr(constraint, "fields", ()):
                groups.append(tuple(constraint.fields))

        deduped_groups = []
        for group in groups:
            if group and group not in deduped_groups:
                deduped_groups.append(group)

        return deduped_groups

    def _get_match_groups(self):
        """Urutan field yang dipakai untuk mendeteksi update data existing"""
        groups = []

        if self.match_fields:
            if isinstance(self.match_fields[0], (list, tuple)):
                groups.extend(tuple(group) for group in self.match_fields)
            else:
                groups.append(tuple(self.match_fields))

        groups.extend(self._get_unique_constraint_groups())

        for field_name in self._get_unique_fields():
            if field_name != 'id':
                groups.append((field_name,))

        deduped_groups = []
        for group in groups:
            if group not in deduped_groups:
                deduped_groups.append(group)

        return deduped_groups

    def _find_existing_instance(self, processed_data):
        """Cari data existing berdasarkan konfigurasi match fields"""
        for group in self._get_match_groups():
            if not all(field_name in processed_data for field_name in group):
                continue

            if not all(
                self._has_matchable_value(field_name, processed_data[field_name])
                for field_name in group
            ):
                continue

            lookup_kwargs = {
                field_name: processed_data[field_name]
                for field_name in group
            }

            instance = self.model.objects.filter(**lookup_kwargs).order_by('pk').first()
            if instance:
                return instance

        return None

    def _normalize_comparison_value(self, value):
        if isinstance(value, models.Model):
            return value.pk
        return value

    def _get_changed_fields(self, existing_instance, processed_data):
        changed_fields = []

        for field_name, new_value in processed_data.items():
            current_value = getattr(existing_instance, field_name, None)

            if self._normalize_comparison_value(current_value) != self._normalize_comparison_value(new_value):
                changed_fields.append(field_name)

        return changed_fields

    def _build_validation_instance(self, processed_data, existing_instance=None):
        instance = existing_instance or self.model()

        for field_name, value in processed_data.items():
            setattr(instance, field_name, value)

        return instance

    def _get_file_extension(self):
        if not self.filename or "." not in self.filename:
            return None
        return f".{self.filename.rsplit('.', 1)[-1].lower()}"

    def _read_binary_content(self):
        if isinstance(self.file_stream, bytes):
            return self.file_stream

        if hasattr(self.file_stream, "seek"):
            self.file_stream.seek(0)

        content = self.file_stream.read()

        if isinstance(content, str):
            return content.encode("utf-8")

        return content

    def _decode_text_content(self):
        raw_bytes = self._read_binary_content()

        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                return raw_bytes.decode(encoding)
            except UnicodeDecodeError:
                continue

        return raw_bytes.decode("utf-8", errors="ignore")

    def _read_spreadsheet_rows(self):
        if isinstance(self.file_stream, bytes):
            book = load_workbook(io.BytesIO(self.file_stream), data_only=True)
        else:
            if hasattr(self.file_stream, "seek"):
                self.file_stream.seek(0)
            book = load_workbook(self.file_stream, data_only=True)

        ws = book.active
        return list(ws.iter_rows(values_only=True))

    def _read_csv_rows(self):
        text_content = self._decode_text_content()

        sample = text_content[:1024]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(io.StringIO(text_content), dialect=dialect)
        return [tuple(row) for row in reader]

    def _rows_to_data(self, rows):
        data = []

        for row_idx, row in enumerate(rows[1:], 2):
            if self.skip_empty_rows and all(cell in (None, "") for cell in row):
                continue

            row_data = {}
            for col_idx, field_name in enumerate(self.columns):
                if col_idx < len(row):
                    row_data[field_name] = row[col_idx]
                else:
                    row_data[field_name] = None

            data.append({
                "row": row_idx,
                "data": row_data,
            })

        return data

    def _resolve_related_lookup_fields(self, related_model):
        priority_names = [
            "nama",
            "name",
            "title",
            "nama_peraturan",
            "nomor_peraturan",
            "lokasi",
            "kota",
            "pangkat",
            "jabatan",
            "eselon",
            "tingkat",
        ]

        resolved_fields = []
        seen_names = set()

        def add_field(field):
            if (
                field.name in seen_names
                or not isinstance(field, (models.CharField, models.TextField))
            ):
                return

            resolved_fields.append(field)
            seen_names.add(field.name)

        for field_name in priority_names:
            try:
                add_field(related_model._meta.get_field(field_name))
            except Exception:
                continue

        for field in related_model._meta.fields:
            if getattr(field, "unique", False):
                add_field(field)

        for field in related_model._meta.fields:
            add_field(field)

        return resolved_fields

    def _resolve_related_instance(self, related_model, value):
        if isinstance(value, models.Model):
            return value

        normalized_value = value.strip() if isinstance(value, str) else value

        if str(normalized_value).isdigit():
            return related_model.objects.get(id=int(normalized_value))

        try:
            return related_model.objects.get(pk=normalized_value)
        except (related_model.DoesNotExist, ValueError, TypeError):
            pass

        if isinstance(normalized_value, str):
            for field in self._resolve_related_lookup_fields(related_model):
                lookup_name = (
                    f"{field.name}__iexact"
                    if isinstance(field, (models.CharField, models.TextField))
                    else field.name
                )

                try:
                    return related_model.objects.get(**{lookup_name: normalized_value})
                except related_model.DoesNotExist:
                    continue
                except related_model.MultipleObjectsReturned:
                    continue

        raise related_model.DoesNotExist
    
    def read_excel(self):
        """Parse Excel file dan return data"""
        self.read_errors = []

        try:
            extension = self._get_file_extension()

            if extension == ".xls":
                raise ValueError(
                    "Format .xls belum didukung. Gunakan file .xlsx atau .csv."
                )

            if extension == ".csv":
                rows = self._read_csv_rows()
            elif extension in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
                rows = self._read_spreadsheet_rows()
            else:
                try:
                    rows = self._read_spreadsheet_rows()
                except Exception:
                    rows = self._read_csv_rows()

            return self._rows_to_data(rows)

        except Exception as e:
            self.read_errors.append(str(e))
            return []
    
    def validate(self, data=None):
        """Validasi data sebelum import"""
        if data is None:
            data = self.read_excel()
        
        self.preview_data = []
        self.errors = []
        self.empty_file_error = False

        if self.read_errors:
            self.errors.extend(self.read_errors)
            return False

        if not data:
            self.empty_file_error = True
            self.errors.append("File tidak berisi data untuk diimport")
            return False
        
        for item in data:
            row_num = item['row']
            row_data = item['data']
            
            try:
                processed_data, errors = self._prepare_processed_data(row_data)
                existing_instance = self._find_existing_instance(processed_data)
                changed_fields = []

                if existing_instance:
                    changed_fields = self._get_changed_fields(existing_instance, processed_data)
                
                # Validasi pakai instance existing agar unique check tidak dianggap create baru
                instance = self._build_validation_instance(processed_data, existing_instance)
                instance.full_clean()
                
                # Check jika ada FK errors
                if errors:
                    error_msg = '; '.join(errors)
                    self.preview_data.append({
                        'row': row_num,
                        'data': row_data,
                        'status': 'error',
                        'message': 'Data tidak valid',
                        'errors': errors
                    })
                    self.errors.append(f"Row {row_num}: {error_msg}")
                else:
                    status = 'new'
                    message = 'Data baru, siap ditambahkan'

                    if existing_instance:
                        if changed_fields:
                            status = 'update'
                            message = f"Data sudah ada, akan diperbarui ({', '.join(changed_fields)})"
                        else:
                            status = 'exists'
                            message = 'Data sudah ada di database, akan dilewati'

                    self.preview_data.append({
                        'row': row_num,
                        'data': row_data,
                        'status': status,
                        'message': message,
                        'errors': []
                    })
            
            except ValidationError as e:
                error_msg = '; '.join([f"{k}: {', '.join(v)}" for k, v in e.message_dict.items()])
                self.preview_data.append({
                    'row': row_num,
                    'data': row_data,
                    'status': 'error',
                    'message': 'Data tidak valid',
                    'errors': [error_msg]
                })
                self.errors.append(f"Row {row_num}: {error_msg}")
            
            except Exception as e:
                self.preview_data.append({
                    'row': row_num,
                    'data': row_data,
                    'status': 'error',
                    'message': 'Terjadi kesalahan saat membaca data',
                    'errors': [str(e)]
                })
                self.errors.append(f"Row {row_num}: {str(e)}")
        
        return len(self.errors) == 0
    
    def import_data(self, data=None):
        """Import data ke database (hanya yang valid)"""
        if data is None:
            data = self.read_excel()

        self.validate(data)

        if self.read_errors or self.empty_file_error:
            return {
                'success': False,
                'imported': 0,
                'updated': 0,
                'skipped': 0,
                'failed': 0,
                'errors': self.errors,
                'preview': self.preview_data
            }
        
        imported = 0
        updated = 0
        skipped = 0
        failed = len([item for item in self.preview_data if item['status'] == 'error'])
        
        try:
            for item in self.preview_data:
                if item['status'] == 'error':
                    continue

                row_data = item['data']
                processed_data, errors = self._prepare_processed_data(row_data)

                if errors:
                    raise ValueError('; '.join(errors))

                existing_instance = self._find_existing_instance(processed_data)

                if item['status'] == 'exists':
                    skipped += 1
                    continue

                if existing_instance:
                    for field_name, value in processed_data.items():
                        setattr(existing_instance, field_name, value)
                    existing_instance.save()
                    updated += 1
                else:
                    self.model.objects.create(**processed_data)
                    imported += 1
            
            return {
                'success': (imported + updated + skipped) > 0 or failed == 0,
                'imported': imported,
                'updated': updated,
                'skipped': skipped,
                'failed': failed,
                'errors': self.errors,
                'preview': self.preview_data
            }
        
        except Exception as e:
            return {
                'success': False,
                'imported': imported,
                'updated': updated,
                'skipped': skipped,
                'failed': failed,
                'errors': [str(e)],
                'preview': self.preview_data
            }
